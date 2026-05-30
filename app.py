from flask import Flask, render_template, request, redirect, url_for, flash, Response, session
from datetime import datetime
from urllib.parse import quote
from werkzeug.security import generate_password_hash, check_password_hash
import json
import os
import csv
import io
import copy
import sqlite3
import tempfile
import openpyxl
from openpyxl.styles import Font
from backup import run_backup, BACKUP_DIR

# Paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
DB_FILE  = os.path.join(DATA_DIR, 'inventory.db')

def _get_or_create_secret_key():
    key_file = os.path.join(DATA_DIR, '.secret_key')
    if os.path.exists(key_file):
        with open(key_file, 'rb') as f:
            return f.read()
    os.makedirs(DATA_DIR, exist_ok=True)
    key = os.urandom(32)
    with open(key_file, 'wb') as f:
        f.write(key)
    return key

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY') or _get_or_create_secret_key()

# Legacy JSON paths — kept only for one-time migration on first run
INVENTORY_FILE    = os.path.join(DATA_DIR, 'inventory.json')
TICKETS_FILE      = os.path.join(DATA_DIR, 'tickets.json')
HISTORY_FILE      = os.path.join(DATA_DIR, 'history.json')
USERS_FILE        = os.path.join(DATA_DIR, 'users.json')
DEVICE_TYPES_FILE = os.path.join(DATA_DIR, 'device_types.json')
LOCATIONS_FILE    = os.path.join(DATA_DIR, 'locations.json')
ISLANDS_FILE      = os.path.join(DATA_DIR, 'islands.json')

# Location codes mapping
LOCATION_CODES = {
    '101': 'Hilo Office – 45 Mohouli St',
    '102': 'Kona Office – 75 Kuakini Hwy',
    '103': 'Honolulu Office – 1001 Bishop St',
    '104': 'Kahului Office – 33 Lono Ave',
    '105': 'Lihue Office – 4444 Rice St',
    '201': 'Remote Office – Big Island',
    '202': 'Remote Office – Oahu',
    '999': 'Mobile/Field Device'
}

LOCATION_ADDRESSES = {
    'Ahukini St':                        '759 Ahukini St, Honolulu, HI 96825',
    'Akahi Lihue admin':                 '3205 Akahi St, Lihue, HI 96766',
    'Apelila':                           '5135 Apelila St, Kapaa, HI 96746',
    'Apuakea Pl':                        '45-634 Apuakea Pl, Kaneohe, HI 96744',
    'Aukoi':                             '2895 Aukoi St, Lihue, HI 96766',
    'Big Island Residential, Inc./Alanoe': '75-5750 Alanoe St, Kailua-Kona 96740',
    'Dominis House':                     '1316 Dominis St, Honolulu, HI 96822',
    'Hale Amau':                         '139 Amau Rd, Hilo, HI 96720',
    'Hale Imua':                         '45-710 Keaahala Rd #42, Kaneohe, HI 96744',
    'Hale Kapili':                       '138 Central Ave, Wailuku, HI 96793',
    'Hale Kokua Kai (Niolo St)':         '91-1001 Niolo St, Ewa Beach, HI 96706',
    'Hale Malie (Punawai St)':           '46-269 Punawai St, Kaneohe, HI 96744',
    'Hale Noho (Awapapa Pl)':            '45-545 Awapapa Pl, Kaneohe, HI 96744',
    'Hilo Case Management':              '190 Keawe St, Ste 212, Hilo, HI 96720',
    'Hilo Hale':                         '208 Wainaku St, Hilo, HI 96720',
    'Iao House':                         '503 Ua Place, Wailuku, HI 96793',
    'Iwilei Admin Office':               '680 Iwilei Rd Ste 600, Honolulu, HI 96817',
    'Iwilei Kauhale':                    '866 Iwilei Rd, Honolulu, HI 96817',
    'Kalani St':                         '75-166 Kalani St Ste 103, Kailua-Kona 96740',
    'Kalani St / ?':                     '75-166 Kalani St Ste 103, Kailua-Kona 96740',
    'Kalani St / Allen':                 '75-166 Kalani St Ste 103, Kailua-Kona 96740',
    'Kalani St / Ann Ho':                '75-166 Kalani St Ste 103, Kailua-Kona 96740',
    'Kalani St / Babbitt':               '75-166 Kalani St Ste 103, Kailua-Kona 96740',
    'Kalani St / Chris':                 '75-166 Kalani St Ste 103, Kailua-Kona 96740',
    'Kalani St / Dr. Bikle':             '75-166 Kalani St Ste 103, Kailua-Kona 96740',
    'Kalani St / Maika':                 '75-166 Kalani St Ste 103, Kailua-Kona 96740',
    'Kalani St / Spare Laptop #1':       '75-166 Kalani St Ste 103, Kailua-Kona 96740',
    'Kalani St / Spare Laptop #2':       '75-166 Kalani St Ste 103, Kailua-Kona 96740',
    'Kalani St / Spare Laptop #3':       '75-166 Kalani St Ste 103, Kailua-Kona 96740',
    "Kalani St/ Kristy's":               '75-166 Kalani St Ste 103, Kailua-Kona 96740',
    'Kalani ST/CM':                      '75-166 Kalani St Ste 103, Kailua-Kona 96740',
    'Kaulana St':                        '425 Kaulana St, Kahului, HI 96732',
    'Kealapua':                          '74-5035 Kealapua St, Kailua-Kona 96740',
    'Keolu Dr':                          '905 Keolu Dr, Kailua, HI 96734',
    'Kinau A':                           '1304 Kinau St, Honolulu, HI 96814',
    'Kinau B':                           '1304 Kinau St, Honolulu, HI 96814',
    'Ko Kakou Hale (Duncan Dr)':         '45-616 Duncan Dr, Kaneohe, HI 96744',
    'Kona Apartments':                   '75-5758 Alahou St, Kailua-Kona 96740',
    'Kona Kokua Housing Inc.':           "75-187 Alaka'i St, Kailua-Kona 96740",
    'Koukalaka':                         '6330 Koukalaka Pl, Kapaa, HI 96746',
    'Maealani Pl':                       '94-243 Maealani Pl, Mililani, HI 96789',
    'Mahani Hale':                       '1731 Mahani Loop, Honolulu, HI 96819',
    'Market Street':                     '105 Market St #102, Wailuku, HI 96793',
    'Maui Kokua':                        '456 S. Lanai St, Kahului, HI 96732',
    'Maui Safe Haven':                   '133 N. Market St, Wailuku, HI 96793',
    'Maunakea Village':                  '1165 Maunakea St, Honolulu, HI 96817',
    'Menehune':                          '718-726 Menehune Lane, Honolulu, HI',
    'Menehune Polly':                    '718-726 Menehune Lane, Honolulu, HI',
    'Menehune RA':                       '718-726 Menehune Lane, Honolulu, HI',
    'Menehune SRA':                      '718-726 Menehune Lane, Honolulu, HI',
    'Pahoa RA':                          '3443 Pahoa Ave, Honolulu, HI 96816',
    'Pahoa SRA':                         '3443 Pahoa Ave, Honolulu, HI 96816',
    'Patch Place':                       '140 & 140A Wainaku Ave, Hilo, HI 96720',
    'Patch Place 8-16 (agency)':         '136 & 136A Wainaku Ave, Hilo, HI 96720',
    'Pearl City Duplex 1':               '1186 Kuokoa St, Pearl City, HI 96782',
    'Punawai Rest Stop (Kuwili)':        '431 Kuwili St, Honolulu, HI 96817',
    'Rooke Ave':                         '2664+A Rooke Ave, Honolulu, HI 96817',
    'Sierra A House':                    '4510 Sierra Drive, Honolulu, HI 96816',
    'Sierra B House':                    '4510 Sierra Drive, Honolulu, HI 96816',
    "Supervisor's Office - Nurse":       '4510 Sierra Drive, Honolulu, HI 96816',
    "Supervisor's Office - Supervisor":  '4510 Sierra Drive, Honolulu, HI 96816',
    'Vineyard':                          '2276 Vineyard St, Wailuku, HI 96793',
    'Wailupe Dr':                        '798 Wailupe Dr, Wailuku, HI 96793',
    'Wellness House, Inc (Ft. Weaver)':  '91-1030 Fort Weaver Rd, Ewa Beach, HI 96706',
}

# Device types and ticket categories
DEVICE_TYPES = ['Laptop', 'Desktop', 'Laptop/Desktop', 'iPhone', 'Android Phone', 'Phone', 'Tablet', 'Monitor', 'Printer', 'Scanner', 'Router', 'Other']

DEVICE_TYPE_ICONS = {
    'Laptop': '💻',
    'Desktop': '🖥️',
    'Laptop/Desktop': '💻',
    'iPhone': '📱',
    'Android Phone': '📱',
    'Phone': '📱',
    'Tablet': '📱',
    'Monitor': '🖥️',
    'Printer': '🖨️',
    'Scanner': '🔍',
    'Router': '📡',
    'Other': '📦',
}

# Maps raw import values (lowercase) to canonical DEVICE_TYPES entries
DEVICE_TYPE_ALIASES = {
    'phone': 'Phone',
    'scanner': 'Scanner',
    'laptop-desktop': 'Laptop/Desktop',
    'laptop desktop': 'Laptop/Desktop',
    'laptop / desktop': 'Laptop/Desktop',
}
TICKET_CATEGORIES = ['Hardware Issue', 'Software Issue', 'Network/WiFi', 'Password Reset', 'Setup/Installation', 'Other']
TICKET_STATUSES = ['Open', 'In Progress', 'Completed', 'Closed']
PRIORITY_LEVELS = ['Low', 'Medium', 'High', 'Urgent']
DEVICE_CONDITIONS = ['New', 'Good', 'Fair', 'Poor', 'Broken']

@app.context_processor
def inject_globals():
    return {'device_type_icons': DEVICE_TYPE_ICONS}

@app.template_filter('urlencode')
def urlencode_filter(s):
    return quote(str(s))

# ── SQLite helpers ────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA journal_mode=WAL')
    conn.execute('PRAGMA foreign_keys=ON')
    return conn

def _get_device(device_id):
    conn = get_db()
    row = conn.execute('SELECT * FROM inventory WHERE id = ?', (device_id,)).fetchone()
    conn.close()
    if not row:
        return None
    d = dict(row)
    d['archived'] = bool(d.get('archived', 0))
    return d

def init_db():
    """Create all tables if they don't exist."""
    conn = get_db()
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS inventory (
            id            INTEGER PRIMARY KEY,
            device_type   TEXT    DEFAULT '',
            model         TEXT    DEFAULT '',
            serial_number TEXT    DEFAULT '',
            assigned_user TEXT    DEFAULT '',
            password      TEXT    DEFAULT '',
            purchase_date TEXT    DEFAULT '',
            condition     TEXT    DEFAULT 'Good',
            location_code TEXT    DEFAULT '',
            island        TEXT    DEFAULT '',
            phone         TEXT    DEFAULT '',
            notes         TEXT    DEFAULT '',
            archived      INTEGER DEFAULT 0,
            archived_at   TEXT    DEFAULT '',
            added_at      TEXT    DEFAULT '',
            updated_at    TEXT    DEFAULT '',
            source        TEXT    DEFAULT 'added'
        );
        CREATE TABLE IF NOT EXISTS tickets (
            id           INTEGER PRIMARY KEY,
            title        TEXT DEFAULT '',
            description  TEXT DEFAULT '',
            category     TEXT DEFAULT '',
            priority     TEXT DEFAULT 'Medium',
            status       TEXT DEFAULT 'Open',
            submitted_by TEXT DEFAULT '',
            created_at   TEXT DEFAULT '',
            updated_at   TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS history (
            id              INTEGER PRIMARY KEY,
            timestamp       TEXT DEFAULT '',
            performed_by    TEXT DEFAULT '',
            action          TEXT DEFAULT '',
            device_id       INTEGER,
            device_type     TEXT DEFAULT '',
            model           TEXT DEFAULT '',
            serial_number   TEXT DEFAULT '',
            location_code   TEXT DEFAULT '',
            island          TEXT DEFAULT '',
            assigned_user   TEXT DEFAULT '',
            changes         TEXT DEFAULT '{}',
            device_snapshot TEXT DEFAULT '{}'
        );
        CREATE TABLE IF NOT EXISTS users (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            name          TEXT UNIQUE NOT NULL,
            password_hash TEXT DEFAULT '',
            is_admin      INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS device_types (
            id   INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL
        );
        CREATE TABLE IF NOT EXISTS locations (
            id     INTEGER PRIMARY KEY AUTOINCREMENT,
            code   TEXT UNIQUE NOT NULL,
            island TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS islands (
            id   INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL
        );
    ''')
    # Add source column to existing databases that don't have it yet
    existing = [row[1] for row in conn.execute('PRAGMA table_info(inventory)').fetchall()]
    if 'source' not in existing:
        conn.execute("ALTER TABLE inventory ADD COLUMN source TEXT DEFAULT 'added'")
    # Add password_hash column to users if missing
    user_cols = [row[1] for row in conn.execute('PRAGMA table_info(users)').fetchall()]
    if 'password_hash' not in user_cols:
        conn.execute("ALTER TABLE users ADD COLUMN password_hash TEXT DEFAULT ''")
    if 'is_admin' not in user_cols:
        conn.execute("ALTER TABLE users ADD COLUMN is_admin INTEGER DEFAULT 0")
    conn.commit()
    conn.close()

def _migrate_json_to_db():
    """One-time migration: import existing JSON data into SQLite."""
    conn = get_db()

    def _json_load(path):
        try:
            with open(path, 'r', encoding='utf-8') as f:
                content = f.read().strip()
                return json.loads(content) if content else []
        except Exception:
            return []

    # Inventory
    if conn.execute('SELECT COUNT(*) FROM inventory').fetchone()[0] == 0:
        for d in _json_load(INVENTORY_FILE):
            conn.execute('''INSERT OR IGNORE INTO inventory
                (id,device_type,model,serial_number,assigned_user,password,
                 purchase_date,condition,location_code,island,phone,notes,
                 archived,archived_at,added_at,updated_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', (
                d.get('id'), d.get('device_type',''), d.get('model',''),
                d.get('serial_number',''), d.get('assigned_user',''),
                d.get('password',''), d.get('purchase_date',''),
                d.get('condition','Good'), d.get('location_code',''),
                d.get('island',''), d.get('phone',''), d.get('notes',''),
                1 if d.get('archived') else 0,
                d.get('archived_at',''), d.get('added_at',''), d.get('updated_at','')))

    # Tickets
    if conn.execute('SELECT COUNT(*) FROM tickets').fetchone()[0] == 0:
        for t in _json_load(TICKETS_FILE):
            conn.execute('''INSERT OR IGNORE INTO tickets
                (id,title,description,category,priority,status,submitted_by,created_at,updated_at)
                VALUES (?,?,?,?,?,?,?,?,?)''', (
                t.get('id'), t.get('title',''), t.get('description',''),
                t.get('category',''), t.get('priority','Medium'),
                t.get('status','Open'), t.get('submitted_by',''),
                t.get('created_at',''), t.get('updated_at','')))

    # History
    if conn.execute('SELECT COUNT(*) FROM history').fetchone()[0] == 0:
        for h in _json_load(HISTORY_FILE):
            conn.execute('''INSERT OR IGNORE INTO history
                (id,timestamp,performed_by,action,device_id,device_type,model,
                 serial_number,location_code,island,assigned_user,changes,device_snapshot)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)''', (
                h.get('id'), h.get('timestamp',''), h.get('performed_by',''),
                h.get('action',''), h.get('device_id'),
                h.get('device_type',''), h.get('model',''),
                h.get('serial_number',''), h.get('location_code',''),
                h.get('island',''), h.get('assigned_user',''),
                json.dumps(h.get('changes',{})),
                json.dumps(h.get('device_snapshot',{}))))

    # Users
    if conn.execute('SELECT COUNT(*) FROM users').fetchone()[0] == 0:
        for name in _json_load(USERS_FILE):
            conn.execute('INSERT OR IGNORE INTO users (name) VALUES (?)', (name,))

    # Device types
    if conn.execute('SELECT COUNT(*) FROM device_types').fetchone()[0] == 0:
        types = _json_load(DEVICE_TYPES_FILE) or DEVICE_TYPES
        for t in types:
            conn.execute('INSERT OR IGNORE INTO device_types (name) VALUES (?)', (t,))

    # Locations
    if conn.execute('SELECT COUNT(*) FROM locations').fetchone()[0] == 0:
        for loc in _json_load(LOCATIONS_FILE):
            conn.execute('INSERT OR IGNORE INTO locations (code,island) VALUES (?,?)',
                         (loc.get('code',''), loc.get('island','')))

    # Islands
    if conn.execute('SELECT COUNT(*) FROM islands').fetchone()[0] == 0:
        for name in _json_load(ISLANDS_FILE):
            conn.execute('INSERT OR IGNORE INTO islands (name) VALUES (?)', (name,))

    conn.commit()
    conn.close()

# Initialize data storage
def init_data_storage():
    os.makedirs(DATA_DIR, exist_ok=True)
    init_db()
    _migrate_json_to_db()

# ── Data access functions ─────────────────────────────────────────────────────

def load_inventory():
    conn = get_db()
    rows = conn.execute('SELECT * FROM inventory').fetchall()
    conn.close()
    result = []
    for row in rows:
        d = dict(row)
        d['archived'] = bool(d.get('archived', 0))
        result.append(d)
    return result

def save_inventory(inventory):
    conn = get_db()
    try:
        conn.execute('BEGIN')
        conn.execute('DELETE FROM inventory')
        for d in inventory:
            conn.execute('''INSERT INTO inventory
                (id,device_type,model,serial_number,assigned_user,password,
                 purchase_date,condition,location_code,island,phone,notes,
                 archived,archived_at,added_at,updated_at,source)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', (
                d.get('id'), d.get('device_type',''), d.get('model',''),
                d.get('serial_number',''), d.get('assigned_user',''),
                d.get('password',''), d.get('purchase_date',''),
                d.get('condition','Good'), d.get('location_code',''),
                d.get('island',''), d.get('phone',''), d.get('notes',''),
                1 if d.get('archived') else 0,
                d.get('archived_at',''), d.get('added_at',''), d.get('updated_at',''),
                d.get('source','added')))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

def load_tickets():
    conn = get_db()
    rows = conn.execute('SELECT * FROM tickets').fetchall()
    conn.close()
    return [dict(row) for row in rows]

def save_tickets(tickets):
    conn = get_db()
    try:
        conn.execute('BEGIN')
        conn.execute('DELETE FROM tickets')
        for t in tickets:
            conn.execute('''INSERT INTO tickets
                (id,title,description,category,priority,status,submitted_by,created_at,updated_at)
                VALUES (?,?,?,?,?,?,?,?,?)''', (
                t.get('id'), t.get('title',''), t.get('description',''),
                t.get('category',''), t.get('priority','Medium'),
                t.get('status','Open'), t.get('submitted_by',''),
                t.get('created_at',''), t.get('updated_at','')))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def load_history():
    conn = get_db()
    rows = conn.execute('SELECT * FROM history ORDER BY id DESC').fetchall()
    conn.close()
    result = []
    for row in rows:
        d = dict(row)
        d['changes'] = json.loads(d.get('changes') or '{}')
        d['device_snapshot'] = json.loads(d.get('device_snapshot') or '{}')
        result.append(d)
    return result

def save_history(history):
    conn = get_db()
    try:
        conn.execute('BEGIN')
        conn.execute('DELETE FROM history')
        for h in history:
            conn.execute('''INSERT INTO history
                (id,timestamp,performed_by,action,device_id,device_type,model,
                 serial_number,location_code,island,assigned_user,changes,device_snapshot)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)''', (
                h.get('id'), h.get('timestamp',''), h.get('performed_by',''),
                h.get('action',''), h.get('device_id'),
                h.get('device_type',''), h.get('model',''),
                h.get('serial_number',''), h.get('location_code',''),
                h.get('island',''), h.get('assigned_user',''),
                json.dumps(h.get('changes',{})),
                json.dumps(h.get('device_snapshot',{}))))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

def log_history(action, device, changes=None):
    conn = get_db()
    conn.execute('''INSERT INTO history
        (timestamp,performed_by,action,device_id,device_type,model,
         serial_number,location_code,island,assigned_user,changes,device_snapshot)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''', (
        datetime.now().strftime('%m-%d-%Y %H:%M:%S'),
        session.get('current_user', 'Unknown') if session else 'System',
        action,
        device.get('id'),
        device.get('device_type', ''),
        device.get('model', ''),
        device.get('serial_number', ''),
        device.get('location_code', ''),
        device.get('island', ''),
        device.get('assigned_user', ''),
        json.dumps(changes or {}),
        json.dumps(copy.deepcopy(device))
    ))
    conn.commit()
    conn.close()

def load_users():
    conn = get_db()
    rows = conn.execute('SELECT name FROM users ORDER BY id').fetchall()
    conn.close()
    return [row['name'] for row in rows]


def load_device_types():
    conn = get_db()
    rows = conn.execute('SELECT name FROM device_types ORDER BY name').fetchall()
    conn.close()
    types = [row['name'] for row in rows]
    if not types:
        save_device_types(DEVICE_TYPES)
        return DEVICE_TYPES[:]
    return types

def save_device_types(types):
    conn = get_db()
    try:
        conn.execute('BEGIN')
        conn.execute('DELETE FROM device_types')
        for t in types:
            conn.execute('INSERT INTO device_types (name) VALUES (?)', (t,))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

def load_locations():
    conn = get_db()
    rows = conn.execute('SELECT code, island FROM locations ORDER BY code').fetchall()
    conn.close()
    return [{'code': row['code'], 'island': row['island']} for row in rows]

def save_locations(locations):
    conn = get_db()
    try:
        conn.execute('BEGIN')
        conn.execute('DELETE FROM locations')
        for loc in locations:
            conn.execute('INSERT INTO locations (code, island) VALUES (?,?)',
                         (loc.get('code',''), loc.get('island','')))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

def load_islands():
    conn = get_db()
    rows = conn.execute('SELECT name FROM islands ORDER BY name').fetchall()
    conn.close()
    return [row['name'] for row in rows]

def save_islands(islands):
    conn = get_db()
    try:
        conn.execute('BEGIN')
        conn.execute('DELETE FROM islands')
        for name in islands:
            conn.execute('INSERT INTO islands (name) VALUES (?)', (name,))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

@app.after_request
def no_cache(response):
    if request.endpoint != 'static':
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

@app.before_request
def require_user():
    allowed = {'select_user', 'logout', 'static'}
    if request.endpoint == 'add_user' and 'current_user' not in session:
        # Allow add_user only when no users exist yet (first-time setup)
        conn = get_db()
        count = conn.execute('SELECT COUNT(*) FROM users').fetchone()[0]
        conn.close()
        if count > 0:
            return redirect(url_for('select_user'))
    elif request.endpoint not in allowed and 'current_user' not in session:
        return redirect(url_for('select_user'))

@app.route('/select-user', methods=['GET', 'POST'])
def select_user():
    conn = get_db()
    all_rows = conn.execute('SELECT name, password_hash, is_admin FROM users ORDER BY id').fetchall()
    conn.close()
    users = [r['name'] for r in all_rows]
    login_users = [r['name'] for r in all_rows if r['password_hash']]
    admin_users = {r['name'] for r in all_rows if r['is_admin']}
    current_is_admin = session.get('current_user') in admin_users

    if request.method == 'POST':
        chosen = request.form.get('user', '').strip()
        password = request.form.get('password', '')
        if not chosen or chosen not in login_users:
            flash('Please select a valid user.', 'error')
            return redirect(url_for('select_user'))
        conn = get_db()
        row = conn.execute('SELECT password_hash FROM users WHERE name = ?', (chosen,)).fetchone()
        conn.close()
        stored_hash = row['password_hash'] if row else ''
        if not check_password_hash(stored_hash, password):
            flash('Incorrect password.', 'error')
            return redirect(url_for('select_user'))
        session['current_user'] = chosen
        session['is_admin'] = chosen in admin_users
        return redirect(url_for('index'))
    return render_template('select_user.html', users=users, login_users=login_users,
                           admin_users=admin_users, current_is_admin=current_is_admin)

@app.route('/add-user', methods=['POST'])
def add_user():
    users = load_users()
    new_name = request.form.get('new_user', '').strip()
    password = request.form.get('new_password', '').strip()
    confirm = request.form.get('confirm_password', '').strip()
    if not new_name:
        flash('Please enter a name.', 'error')
    elif not password:
        flash('Please enter a password.', 'error')
    elif password != confirm:
        flash('Passwords do not match.', 'error')
    elif new_name in users:
        flash(f'"{new_name}" already exists.', 'error')
    else:
        conn = get_db()
        try:
            count = conn.execute('SELECT COUNT(*) FROM users').fetchone()[0]
            make_admin = 1 if count == 0 else 0
            conn.execute('INSERT INTO users (name, password_hash, is_admin) VALUES (?, ?, ?)',
                         (new_name, generate_password_hash(password), make_admin))
            conn.commit()
            if make_admin:
                flash(f'User "{new_name}" added as admin (first account).', 'success')
            else:
                flash(f'User "{new_name}" added.', 'success')
        except sqlite3.IntegrityError:
            conn.rollback()
            flash(f'"{new_name}" already exists.', 'error')
        finally:
            conn.close()
    return redirect(url_for('select_user'))

@app.route('/delete-user', methods=['POST'])
def delete_user():
    users = load_users()
    name = request.form.get('user', '').strip()
    password = request.form.get('password', '')
    if name not in users:
        flash('User not found.', 'error')
        return redirect(url_for('select_user'))
    conn = get_db()
    row = conn.execute('SELECT password_hash FROM users WHERE name = ?', (name,)).fetchone()
    stored_hash = row['password_hash'] if row else ''
    if not stored_hash or not check_password_hash(stored_hash, password):
        conn.close()
        flash('Incorrect password. User not removed.', 'error')
        return redirect(url_for('select_user'))
    conn.execute('DELETE FROM users WHERE name = ?', (name,))
    conn.commit()
    conn.close()
    if session.get('current_user') == name:
        session.pop('current_user', None)
    flash(f'User "{name}" removed.', 'success')
    return redirect(url_for('select_user'))

@app.route('/change-password', methods=['POST'])
def change_password():
    name = request.form.get('user', '').strip()
    current = request.form.get('current_password', '')
    new_pw = request.form.get('new_password', '').strip()
    confirm = request.form.get('confirm_password', '').strip()
    is_admin_reset = request.form.get('admin_reset') == '1'

    if not name:
        flash('Please select a user.', 'error')
        return redirect(url_for('select_user'))

    # Check if logged-in user is admin
    conn = get_db()
    actor = session.get('current_user', '')
    actor_row = conn.execute('SELECT is_admin FROM users WHERE name = ?', (actor,)).fetchone()
    actor_is_admin = bool(actor_row and actor_row['is_admin'])

    row = conn.execute('SELECT password_hash FROM users WHERE name = ?', (name,)).fetchone()
    if not row:
        conn.close()
        flash('User not found.', 'error')
        return redirect(url_for('select_user'))

    # Reject admin_reset attempt from non-admins early with a clear message
    if is_admin_reset and not actor_is_admin:
        conn.close()
        flash('Only admins can reset passwords.', 'error')
        return redirect(url_for('select_user'))

    # Admin reset: skip current password check
    if is_admin_reset and actor_is_admin:
        if not new_pw:
            conn.close()
            flash('New password cannot be empty.', 'error')
            return redirect(url_for('select_user'))
        if new_pw != confirm:
            conn.close()
            flash('New passwords do not match.', 'error')
            return redirect(url_for('select_user'))
        conn.execute('UPDATE users SET password_hash = ? WHERE name = ?',
                     (generate_password_hash(new_pw), name))
        conn.commit()
        conn.close()
        flash(f'Password reset for "{name}".', 'success')
        return redirect(url_for('select_user'))

    # Normal change: requires current password
    if not check_password_hash(row['password_hash'], current):
        conn.close()
        flash('Current password is incorrect.', 'error')
        return redirect(url_for('select_user'))
    if not new_pw:
        conn.close()
        flash('New password cannot be empty.', 'error')
        return redirect(url_for('select_user'))
    if new_pw == current:
        conn.close()
        flash('New password must be different from your current password.', 'error')
        return redirect(url_for('select_user'))
    if new_pw != confirm:
        conn.close()
        flash('New passwords do not match.', 'error')
        return redirect(url_for('select_user'))
    conn.execute('UPDATE users SET password_hash = ? WHERE name = ?',
                 (generate_password_hash(new_pw), name))
    conn.commit()
    conn.close()
    flash(f'Password updated for "{name}".', 'success')
    return redirect(url_for('select_user'))

@app.route('/set-admin', methods=['POST'])
def set_admin():
    actor = session.get('current_user', '')
    conn = get_db()
    actor_row = conn.execute('SELECT is_admin FROM users WHERE name = ?', (actor,)).fetchone()
    if not actor_row or not actor_row['is_admin']:
        conn.close()
        flash('Only admins can change admin status.', 'error')
        return redirect(url_for('select_user'))
    name = request.form.get('user', '').strip()
    try:
        make_admin = int(request.form.get('is_admin', '0'))
        if make_admin not in (0, 1):
            raise ValueError
    except ValueError:
        conn.close()
        flash('Invalid request.', 'error')
        return redirect(url_for('select_user'))
    if not name:
        conn.close()
        flash('Invalid request.', 'error')
        return redirect(url_for('select_user'))
    # Prevent removing own admin if last admin — wrap in transaction to avoid TOCTOU
    conn.execute('BEGIN IMMEDIATE')
    if name == actor and make_admin == 0:
        admin_count = conn.execute('SELECT COUNT(*) FROM users WHERE is_admin = 1').fetchone()[0]
        if admin_count <= 1:
            conn.rollback()
            conn.close()
            flash('Cannot remove admin — you are the only admin.', 'error')
            return redirect(url_for('select_user'))
    result = conn.execute('UPDATE users SET is_admin = ? WHERE name = ?', (make_admin, name))
    if result.rowcount == 0:
        conn.rollback()
        conn.close()
        flash('User not found.', 'error')
        return redirect(url_for('select_user'))
    conn.commit()
    conn.close()
    if name == actor:
        session['is_admin'] = bool(make_admin)
    label = 'granted admin to' if make_admin else 'removed admin from'
    flash(f'Successfully {label} "{name}".', 'success')
    return redirect(url_for('select_user'))

@app.route('/force-backup', methods=['POST'])
def force_backup():
    actor = session.get('current_user', '')
    conn = get_db()
    row = conn.execute('SELECT is_admin FROM users WHERE name = ?', (actor,)).fetchone()
    conn.close()
    if not row or not row['is_admin']:
        flash('Only admins can run a manual backup.', 'error')
        return redirect(request.referrer or url_for('index'))
    try:
        run_backup()
        flash(f'Backup saved to {BACKUP_DIR}', 'success')
    except Exception as e:
        flash(f'Backup failed: {e}', 'error')
    return redirect(request.referrer or url_for('index'))

@app.route('/logout')
def logout():
    session.pop('current_user', None)
    session.pop('is_admin', None)
    return redirect(url_for('select_user'))

# Routes
@app.route('/')
def index():
    """Dashboard homepage"""
    inventory = load_inventory()
    tickets = load_tickets()

    selected_island = request.args.get('island', '')
    archive_filter = request.args.get('archive_filter', 'active')

    # Build island list for filter dropdown
    all_islands = _get_all_islands(inventory)

    # Filter by archive status
    if archive_filter == 'archived':
        filtered = [d for d in inventory if d.get('archived')]
    elif archive_filter == 'all':
        filtered = inventory[:]
    else:
        filtered = [d for d in inventory if not d.get('archived')]

    # Filter by island if selected
    if selected_island:
        filtered = [d for d in filtered if d.get('island', '').strip() == selected_island]

    # Calculate statistics
    total_devices = len(filtered)

    # Devices by location (sorted alphabetically)
    devices_by_location_raw = {}
    for device in filtered:
        location = device.get('location_code', 'Unknown')
        devices_by_location_raw[location] = devices_by_location_raw.get(location, 0) + 1
    devices_by_location = dict(sorted(devices_by_location_raw.items(), key=lambda x: x[0].lower()))

    # Open tickets
    open_tickets = [t for t in tickets if t.get('status') in ['Open', 'In Progress']]
    total_open_tickets = len(open_tickets)

    # Tickets by category
    tickets_by_category = {}
    for ticket in tickets:
        category = ticket.get('category', 'Other')
        tickets_by_category[category] = tickets_by_category.get(category, 0) + 1

    # Recent tickets (last 5)
    recent_tickets = sorted(tickets, key=lambda x: _parse_ts(x.get('created_at', '')), reverse=True)[:5]

    return render_template('index.html',
                         total_devices=total_devices,
                         devices_by_location=devices_by_location,
                         location_addresses=LOCATION_ADDRESSES,
                         total_open_tickets=total_open_tickets,
                         all_islands=all_islands,
                         selected_island=selected_island,
                         archive_filter=archive_filter,
                         tickets_by_category=tickets_by_category,
                         recent_tickets=recent_tickets)

@app.route('/islands/add', methods=['POST'])
def add_island():
    islands = load_islands()
    name = request.form.get('new_island', '').strip()
    if not name:
        flash('Please enter an island name.', 'error')
    elif name in islands:
        flash(f'"{name}" already exists.', 'error')
    else:
        islands.append(name)
        save_islands(islands)
        log_history('Island Added', {
            'id': None, 'device_type': '', 'model': f'Island "{name}"',
            'serial_number': '', 'location_code': '', 'island': name, 'assigned_user': ''
        })
        flash(f'Island "{name}" added.', 'success')
    return redirect(url_for('inventory_list'))

@app.route('/islands/delete', methods=['POST'])
def delete_island():
    islands = load_islands()
    name = request.form.get('island', '').strip()
    confirm = request.form.get('confirm_delete', 'false')
    if name in islands:
        inventory = load_inventory()
        in_use = sum(1 for d in inventory if d.get('island', '') == name)
        if in_use > 0 and confirm != 'true':
            flash(f'WARNING: {in_use} device(s) are assigned to island "{name}". '
                  f'Removing it won\'t change those devices, but "{name}" will disappear from filters. '
                  f'Click Remove again to confirm.', 'error')
            return redirect(url_for('inventory_list'))
        islands.remove(name)
        save_islands(islands)
        log_history('Island Deleted', {
            'id': None, 'device_type': '', 'model': f'Island "{name}"',
            'serial_number': '', 'location_code': '', 'island': name, 'assigned_user': ''
        })
        flash(f'Island "{name}" removed.', 'success')
    else:
        flash('Island not found.', 'error')
    return redirect(url_for('inventory_list'))

@app.route('/locations/add', methods=['POST'])
def add_location():
    locations = load_locations()
    code = request.form.get('new_location', '').strip()
    island = request.form.get('location_island', '').strip()
    if not code:
        flash('Please enter a location name.', 'error')
    elif any(l['code'] == code for l in locations):
        flash(f'"{code}" already exists.', 'error')
    else:
        locations.append({'code': code, 'island': island})
        save_locations(locations)
        log_history('Location Added', {
            'id': None, 'device_type': '', 'model': f'Location "{code}" (Island: {island or "—"})',
            'serial_number': '', 'location_code': code, 'island': island, 'assigned_user': ''
        })
        flash(f'Location "{code}" added.', 'success')
    return redirect(url_for('inventory_list'))

@app.route('/locations/delete', methods=['POST'])
def delete_location():
    locations = load_locations()
    code = request.form.get('location', '').strip()
    if not any(l['code'] == code for l in locations):
        flash('Location not found.', 'error')
        return redirect(url_for('inventory_list'))
    inventory = load_inventory()
    in_use = sum(1 for d in inventory if d.get('location_code', '') == code)
    confirm = request.form.get('confirm_delete', 'false')
    if in_use > 0 and confirm != 'true':
        flash(f'WARNING: {in_use} device(s) use location "{code}". Submit again to confirm removal.', 'error')
        return redirect(url_for('inventory_list'))
    locations = [l for l in locations if l['code'] != code]
    save_locations(locations)
    log_history('Location Deleted', {
        'id': None, 'device_type': '', 'model': f'Location "{code}"',
        'serial_number': '', 'location_code': code, 'island': '', 'assigned_user': ''
    })
    flash(f'Location "{code}" removed.', 'success')
    return redirect(url_for('inventory_list'))

@app.route('/device-types/add', methods=['POST'])
def add_device_type():
    types = load_device_types()
    new_type = request.form.get('new_type', '').strip()
    if not new_type:
        flash('Please enter a device type name.', 'error')
    elif new_type in types:
        flash(f'"{new_type}" already exists.', 'error')
    else:
        types.append(new_type)
        save_device_types(types)
        log_history('Device Type Added', {
            'id': None, 'device_type': new_type, 'model': f'Device type "{new_type}"',
            'serial_number': '', 'location_code': '', 'island': '', 'assigned_user': ''
        })
        flash(f'Device type "{new_type}" added.', 'success')
    return redirect(url_for('inventory_list'))

@app.route('/device-types/delete', methods=['POST'])
def delete_device_type():
    types = load_device_types()
    name = request.form.get('type', '').strip()
    confirm = request.form.get('confirm_delete', 'false')
    if name in types:
        inventory = load_inventory()
        in_use = sum(1 for d in inventory if d.get('device_type', '') == name)
        if in_use > 0 and confirm != 'true':
            flash(f'WARNING: {in_use} device(s) are currently using type "{name}". '
                  f'The type will be removed from the dropdown but those devices will keep their existing type label. '
                  f'Click Remove again to confirm.', 'error')
            return redirect(url_for('inventory_list') + '#deviceTypesModal')
        types.remove(name)
        save_device_types(types)
        log_history('Device Type Deleted', {
            'id': None, 'device_type': name, 'model': f'Device type "{name}"',
            'serial_number': '', 'location_code': '', 'island': '', 'assigned_user': ''
        })
        flash(f'Device type "{name}" removed.', 'success')
    else:
        flash('Device type not found.', 'error')
    return redirect(url_for('inventory_list'))

@app.route('/inventory')
def inventory_list():
    """Display all inventory items"""
    inventory = load_inventory()

    # Get filter parameters
    search = request.args.get('search', '').lower()
    device_type = request.args.get('device_type', '')
    location = request.args.get('location', '')
    selected_island = request.args.get('island', '')
    archive_filter = request.args.get('archive_filter', 'active')  # active | archived | all

    # Filter by archive status first
    if archive_filter == 'archived':
        filtered_inventory = [item for item in inventory if item.get('archived')]
    elif archive_filter == 'all':
        filtered_inventory = inventory[:]
    else:
        filtered_inventory = [item for item in inventory if not item.get('archived')]

    if selected_island:
        filtered_inventory = [item for item in filtered_inventory if item.get('island', '').strip() == selected_island]

    if search:
        filtered_inventory = [
            item for item in filtered_inventory
            if search in item.get('model', '').lower() or
               search in item.get('serial_number', '').lower() or
               search in item.get('assigned_user', '').lower() or
               search in item.get('location_code', '').lower()
        ]

    if device_type:
        filtered_inventory = [item for item in filtered_inventory if item.get('device_type') == device_type]

    if location:
        filtered_inventory = [item for item in filtered_inventory if item.get('location_code') == location]

    # Add location names to items
    for item in filtered_inventory:
        item['location_name'] = LOCATION_CODES.get(item.get('location_code', ''), item.get('location_code', 'Unknown'))

    locations_by_island = _build_locations_by_island(inventory)
    all_islands = _get_all_islands(inventory)

    # Build device type list from saved types + any in inventory
    saved_types = load_device_types()
    all_types_in_inventory = sorted({item.get('device_type', '').strip() for item in inventory if item.get('device_type', '').strip()})
    device_types_for_filter = sorted(set(saved_types) | set(all_types_in_inventory))

    saved_locations = load_locations()
    saved_islands = sorted(load_islands())

    return render_template('inventory.html',
                         inventory=filtered_inventory,
                         full_inventory=inventory,
                         device_types=device_types_for_filter,
                         saved_device_types=sorted(saved_types),
                         saved_locations=saved_locations,
                         saved_islands=saved_islands,
                         location_codes=LOCATION_CODES,
                         locations_by_island=locations_by_island,
                         all_islands=all_islands,
                         search=search,
                         selected_device_type=device_type,
                         selected_location=location,
                         selected_island=selected_island,
                         archive_filter=archive_filter)

@app.route('/inventory/add', methods=['GET', 'POST'])
def add_device():
    """Add a new device to inventory"""
    if request.method == 'POST':
        # Validate required fields
        device_type = request.form.get('device_type')
        model = request.form.get('model')
        serial_number = request.form.get('serial_number')
        location_code = request.form.get('location_code')
        
        if not all([device_type, model, serial_number, location_code]):
            flash('Please fill in all required fields.', 'error')
            return redirect(url_for('add_device'))

        now = datetime.now().strftime('%m-%d-%Y %H:%M:%S')
        new_device = {
            'device_type': device_type,
            'model': model,
            'serial_number': serial_number,
            'assigned_user': request.form.get('assigned_user', ''),
            'password': request.form.get('password', ''),
            'purchase_date': _normalize_date(request.form.get('purchase_date', '')),
            'condition': request.form.get('condition', 'Good'),
            'location_code': location_code,
            'island': _island_for_location(location_code),
            'phone': request.form.get('phone', ''),
            'notes': request.form.get('notes', ''),
            'added_at': now,
            'updated_at': '',
            'source': 'added'
        }
        conn = get_db()
        cur = conn.execute('''INSERT INTO inventory
            (device_type, model, serial_number, assigned_user, password,
             purchase_date, condition, location_code, island, phone, notes,
             archived, added_at, updated_at, source)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,0,?,?,?)''', (
            new_device['device_type'], new_device['model'], new_device['serial_number'],
            new_device['assigned_user'], new_device['password'], new_device['purchase_date'],
            new_device['condition'], new_device['location_code'], new_device['island'],
            new_device['phone'], new_device['notes'],
            new_device['added_at'], new_device['updated_at'], new_device['source']
        ))
        new_device['id'] = cur.lastrowid
        conn.commit()
        conn.close()
        log_history('Added', new_device)
        flash(f'Device {model} added successfully!', 'success')
        return redirect(url_for('inventory_list'))
    
    inventory = load_inventory()
    all_islands = _get_all_islands(inventory)
    return render_template('add_device.html',
                         device_types=load_device_types(),
                         locations_by_island=_build_locations_by_island(inventory),
                         all_islands=all_islands,
                         location_codes=LOCATION_CODES,
                         conditions=DEVICE_CONDITIONS)

@app.route('/inventory/edit/<int:device_id>', methods=['GET', 'POST'])
def edit_device(device_id):
    """Edit an existing device"""
    device = _get_device(device_id)
    if not device:
        flash('Device not found.', 'error')
        return redirect(url_for('inventory_list'))

    if request.method == 'POST':
        if not all([request.form.get('device_type'), request.form.get('model'),
                    request.form.get('serial_number'), request.form.get('location_code')]):
            flash('Please fill in all required fields.', 'error')
            return redirect(url_for('edit_device', device_id=device_id))
        track_fields = ['device_type', 'model', 'serial_number', 'assigned_user', 'password', 'location_code', 'phone', 'purchase_date', 'condition', 'notes']
        changes = {f: {'from': device.get(f, ''), 'to': request.form.get(f, device.get(f, ''))} for f in track_fields if str(device.get(f, '')) != str(request.form.get(f, device.get(f, '')))}
        old_island = device.get('island', '')
        new_island = _island_for_location(request.form.get('location_code', ''))
        if old_island != new_island:
            changes['island'] = {'from': old_island, 'to': new_island}
        condition = request.form.get('condition', 'Good')
        device.update({
            'device_type': request.form.get('device_type'),
            'model': request.form.get('model'),
            'serial_number': request.form.get('serial_number'),
            'assigned_user': request.form.get('assigned_user', ''),
            'password': request.form.get('password', ''),
            'purchase_date': _normalize_date(request.form.get('purchase_date', '')),
            'condition': condition if condition in DEVICE_CONDITIONS else 'Good',
            'location_code': request.form.get('location_code'),
            'island': new_island,
            'phone': request.form.get('phone', ''),
            'notes': request.form.get('notes', ''),
            'updated_at': datetime.now().strftime('%m-%d-%Y %H:%M:%S'),
        })
        conn = get_db()
        conn.execute('''UPDATE inventory SET
            device_type=?, model=?, serial_number=?, assigned_user=?, password=?,
            purchase_date=?, condition=?, location_code=?, island=?, phone=?, notes=?,
            updated_at=? WHERE id=?''', (
            device['device_type'], device['model'], device['serial_number'],
            device['assigned_user'], device['password'], device['purchase_date'],
            device['condition'], device['location_code'], device['island'],
            device['phone'], device['notes'], device['updated_at'], device_id
        ))
        conn.commit()
        conn.close()
        log_history('Edited', device, changes)
        flash(f'Device {device["model"]} updated successfully!', 'success')
        return redirect(url_for('inventory_list'))

    inventory = load_inventory()
    all_islands = _get_all_islands(inventory)
    device['location_name'] = LOCATION_CODES.get(device.get('location_code', ''), device.get('location_code', 'Unknown'))
    return render_template('edit_device.html',
                         device=device,
                         device_types=load_device_types(),
                         locations_by_island=_build_locations_by_island(inventory),
                         all_islands=all_islands,
                         location_codes=LOCATION_CODES,
                         conditions=DEVICE_CONDITIONS)

@app.route('/inventory/delete/<int:device_id>', methods=['POST'])
def delete_device(device_id):
    """Delete a device from inventory"""
    device = _get_device(device_id)
    if not device:
        flash('Device not found.', 'error')
        return redirect(url_for('inventory_list'))
    log_history('Deleted', device)
    conn = get_db()
    conn.execute('DELETE FROM inventory WHERE id=?', (device_id,))
    conn.commit()
    conn.close()
    flash('Device deleted successfully!', 'success')
    return redirect(url_for('inventory_list'))

@app.route('/inventory/archive/<int:device_id>', methods=['POST'])
def archive_device(device_id):
    """Archive a device (mark as predecessor/retired)"""
    device = _get_device(device_id)
    if device:
        now = datetime.now().strftime('%m-%d-%Y %H:%M:%S')
        device['archived'] = True
        device['archived_at'] = now
        conn = get_db()
        try:
            conn.execute('BEGIN')
            conn.execute('UPDATE inventory SET archived=1, archived_at=? WHERE id=?', (now, device_id))
            conn.execute('''INSERT INTO history
                (timestamp,performed_by,action,device_id,device_type,model,
                 serial_number,location_code,island,assigned_user,changes,device_snapshot)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''', (
                now,
                session.get('current_user', 'Unknown'),
                'Archived',
                device.get('id'), device.get('device_type', ''), device.get('model', ''),
                device.get('serial_number', ''), device.get('location_code', ''),
                device.get('island', ''), device.get('assigned_user', ''),
                '{}', json.dumps(copy.deepcopy(device))
            ))
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()
        flash(f'Device {device["model"]} archived.', 'success')
    else:
        flash('Device not found.', 'error')
    return redirect(request.referrer or url_for('inventory_list'))

@app.route('/inventory/mark-tbd/<int:device_id>', methods=['POST'])
def mark_tbd(device_id):
    """Keep device active but set assigned user to TBD"""
    device = _get_device(device_id)
    if device:
        now = datetime.now().strftime('%m-%d-%Y %H:%M:%S')
        conn = get_db()
        try:
            conn.execute('BEGIN')
            conn.execute('UPDATE inventory SET archived=1, archived_at=? WHERE id=?', (now, device_id))
            cur = conn.execute('''INSERT INTO inventory
                (device_type, model, serial_number, assigned_user, password,
                 purchase_date, condition, location_code, island, phone, notes,
                 archived, added_at, updated_at, source)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,0,?,?,?)''', (
                device['device_type'], device['model'], device['serial_number'],
                'TBD', 'TBD', device['purchase_date'], device['condition'],
                device['location_code'], device['island'], 'TBD', device['notes'],
                now, '', device.get('source', 'added')
            ))
            new_id = cur.lastrowid
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()
        device['archived'] = True
        device['archived_at'] = now
        log_history('Archived (TBD copy created)', device)
        new_device = dict(device)
        new_device['id'] = new_id
        new_device['assigned_user'] = 'TBD'
        new_device['password'] = 'TBD'
        new_device['phone'] = 'TBD'
        new_device['archived'] = False
        new_device.pop('archived_at', None)
        new_device['added_at'] = now
        new_device['updated_at'] = ''
        log_history('Added (TBD copy)', new_device)
        flash(f'Original {device["model"]} archived. Active copy created with TBD fields.', 'success')
    else:
        flash('Device not found.', 'error')
    return redirect(request.referrer or url_for('inventory_list'))

@app.route('/inventory/unarchive/<int:device_id>', methods=['POST'])
def unarchive_device(device_id):
    """Restore an archived device to active inventory"""
    device = _get_device(device_id)
    if device:
        conn = get_db()
        conn.execute('UPDATE inventory SET archived=0, archived_at=? WHERE id=?', ('', device_id))
        conn.commit()
        conn.close()
        device['archived'] = False
        device.pop('archived_at', None)
        log_history('Restored', device)
        flash(f'Device {device["model"]} restored to active inventory.', 'success')
    else:
        flash('Device not found.', 'error')
    return redirect(request.referrer or url_for('inventory_list'))

@app.route('/inventory/history/undo/<int:history_id>', methods=['POST'])
def undo_history(history_id):
    conn = get_db()
    row = conn.execute('SELECT * FROM history WHERE id=?', (history_id,)).fetchone()
    conn.close()
    if not row:
        flash('History entry not found.', 'error')
        return redirect(url_for('inventory_history'))

    entry = dict(row)
    entry['changes'] = json.loads(entry.get('changes') or '{}')
    entry['device_snapshot'] = json.loads(entry.get('device_snapshot') or '{}')

    action = entry.get('action', '')
    device_id = entry.get('device_id')

    # --- TBD pair: undo both archive + TBD copy together ---
    if action in ('Archived (TBD copy created)', 'Added (TBD copy)'):
        conn = get_db()
        pair_rows = conn.execute(
            '''SELECT * FROM history WHERE timestamp=? AND performed_by=?
               AND action IN ('Archived (TBD copy created)', 'Added (TBD copy)')''',
            (entry.get('timestamp'), entry.get('performed_by'))
        ).fetchall()
        conn.close()
        pair = []
        for r in pair_rows:
            d = dict(r)
            d['changes'] = json.loads(d.get('changes') or '{}')
            d['device_snapshot'] = json.loads(d.get('device_snapshot') or '{}')
            pair.append(d)

        archive_entry = next((e for e in pair if e['action'] == 'Archived (TBD copy created)'), None)
        tbd_entry     = next((e for e in pair if e['action'] == 'Added (TBD copy)'), None)

        if archive_entry and tbd_entry:
            conn = get_db()
            try:
                conn.execute('BEGIN')
                conn.execute('UPDATE inventory SET archived=0, archived_at=? WHERE id=?',
                             ('', archive_entry['device_id']))
                conn.execute('DELETE FROM inventory WHERE id=?', (tbd_entry['device_id'],))
                conn.execute('DELETE FROM history WHERE id IN (?,?)',
                             (archive_entry['id'], tbd_entry['id']))
                conn.commit()
            except Exception:
                conn.rollback()
                raise
            finally:
                conn.close()
            log_history('Undone: Archive+TBD', {
                'id': archive_entry['device_id'],
                'device_type': archive_entry.get('device_type', ''),
                'model': archive_entry.get('model', ''),
                'serial_number': archive_entry.get('serial_number', ''),
                'location_code': archive_entry.get('location_code', ''),
                'island': archive_entry.get('island', ''),
                'assigned_user': archive_entry.get('assigned_user', ''),
            })
            flash(f'Undone: archive and TBD copy for {archive_entry.get("model")} reversed.', 'success')
        else:
            flash('Could not find matching TBD pair to undo.', 'error')

    # --- Added: delete the device ---
    elif action == 'Added':
        conn = get_db()
        conn.execute('DELETE FROM inventory WHERE id=?', (device_id,))
        conn.execute('DELETE FROM history WHERE id=?', (history_id,))
        conn.commit()
        conn.close()
        log_history('Undone: Added', {
            'id': device_id, 'device_type': entry.get('device_type', ''),
            'model': entry.get('model', ''), 'serial_number': entry.get('serial_number', ''),
            'location_code': entry.get('location_code', ''), 'island': entry.get('island', ''),
            'assigned_user': entry.get('assigned_user', ''),
        })
        flash(f'Undone: removed added device {entry.get("model")}.', 'success')

    # --- Edited: revert fields to their previous values ---
    elif action == 'Edited':
        changes = entry.get('changes', {})
        if device_id and changes:
            EDITABLE_SQL = {
                'device_type':   'device_type = ?',
                'model':         'model = ?',
                'serial_number': 'serial_number = ?',
                'assigned_user': 'assigned_user = ?',
                'password':      'password = ?',
                'location_code': 'location_code = ?',
                'phone':         'phone = ?',
                'purchase_date': 'purchase_date = ?',
                'condition':     'condition = ?',
                'notes':         'notes = ?',
                'island':        'island = ?',
            }
            set_parts, values = [], []
            for field, diff in changes.items():
                if field in EDITABLE_SQL and 'from' in diff:
                    set_parts.append(EDITABLE_SQL[field])
                    values.append(diff['from'])
            if not set_parts:
                flash('No editable fields to revert.', 'error')
            else:
                values.append(device_id)
                conn = get_db()
                conn.execute('UPDATE inventory SET ' + ', '.join(set_parts) + ' WHERE id=?', values)
                conn.execute('DELETE FROM history WHERE id=?', (history_id,))
                conn.commit()
                conn.close()
                device = _get_device(device_id)
                log_history('Undone: Edited', device or {'id': device_id, 'model': entry.get('model', '')})
                flash(f'Undone: reverted edit on {entry.get("model")}.', 'success')
        else:
            flash('No changes to revert or device not found.', 'error')

    # --- Deleted: restore from snapshot ---
    elif action == 'Deleted':
        snapshot = entry.get('device_snapshot', {})
        if snapshot:
            conn = get_db()
            try:
                conn.execute('BEGIN')
                cur = conn.execute('''INSERT INTO inventory
                    (device_type, model, serial_number, assigned_user, password,
                     purchase_date, condition, location_code, island, phone, notes,
                     archived, archived_at, added_at, updated_at, source)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', (
                    snapshot.get('device_type', ''), snapshot.get('model', ''),
                    snapshot.get('serial_number', ''), snapshot.get('assigned_user', ''),
                    snapshot.get('password', ''), snapshot.get('purchase_date', ''),
                    snapshot.get('condition', 'Good'), snapshot.get('location_code', ''),
                    snapshot.get('island', ''), snapshot.get('phone', ''), snapshot.get('notes', ''),
                    1 if snapshot.get('archived') else 0,
                    snapshot.get('archived_at', ''), snapshot.get('added_at', ''),
                    snapshot.get('updated_at', ''), snapshot.get('source', 'added')
                ))
                new_id = cur.lastrowid
                conn.execute('DELETE FROM history WHERE id=?', (history_id,))
                conn.commit()
            except Exception:
                conn.rollback()
                raise
            finally:
                conn.close()
            snapshot['id'] = new_id
            log_history('Undone: Deleted', snapshot)
            flash(f'Undone: restored deleted device {entry.get("model")}.', 'success')
        else:
            flash('No snapshot available — this entry was created before undo support was added.', 'error')

    # --- Archived: restore to active ---
    elif action == 'Archived':
        device = _get_device(device_id)
        if device:
            conn = get_db()
            conn.execute('UPDATE inventory SET archived=0, archived_at=? WHERE id=?', ('', device_id))
            conn.execute('DELETE FROM history WHERE id=?', (history_id,))
            conn.commit()
            conn.close()
            device['archived'] = False
            log_history('Undone: Archived', device)
            flash(f'Undone: {entry.get("model")} restored to active inventory.', 'success')
        else:
            flash('Device not found in inventory.', 'error')

    # --- Restored: re-archive ---
    elif action == 'Restored':
        device = _get_device(device_id)
        if device:
            now = datetime.now().strftime('%m-%d-%Y %H:%M:%S')
            conn = get_db()
            conn.execute('UPDATE inventory SET archived=1, archived_at=? WHERE id=?', (now, device_id))
            conn.execute('DELETE FROM history WHERE id=?', (history_id,))
            conn.commit()
            conn.close()
            device['archived'] = True
            device['archived_at'] = now
            log_history('Undone: Restored', device)
            flash(f'Undone: {entry.get("model")} re-archived.', 'success')
        else:
            flash('Device not found.', 'error')

    else:
        flash(f'Undo not supported for action: {action}', 'error')

    return redirect(url_for('inventory_history'))


@app.route('/inventory/history')
def inventory_history():
    """Display inventory change history"""
    all_history = load_history()
    search = request.args.get('search', '').lower()
    action_filter = request.args.get('action', '')
    history = all_history
    if search:
        history = [e for e in history if search in e.get('model', '').lower() or
                   search in e.get('serial_number', '').lower() or
                   search in e.get('location_code', '').lower() or
                   search in e.get('assigned_user', '').lower()]
    if action_filter:
        history = [e for e in history if e.get('action', '') == action_filter]
    all_actions = sorted({e.get('action', '') for e in all_history})
    return render_template('inventory_history.html',
                           history=history,
                           all_actions=all_actions,
                           search=search,
                           selected_action=action_filter)

@app.route('/tickets')
def tickets_list():
    """Display all tickets"""
    tickets = load_tickets()
    
    # Get filter parameters
    search = request.args.get('search', '').lower()
    status = request.args.get('status', '')
    category = request.args.get('category', '')
    priority = request.args.get('priority', '')
    
    # Filter tickets
    filtered_tickets = tickets
    
    if search:
        filtered_tickets = [
            ticket for ticket in filtered_tickets
            if search in ticket.get('title', '').lower() or
               search in ticket.get('description', '').lower() or
               search in ticket.get('submitted_by', '').lower()
        ]
    
    if status:
        filtered_tickets = [ticket for ticket in filtered_tickets if ticket.get('status') == status]
    
    if category:
        filtered_tickets = [ticket for ticket in filtered_tickets if ticket.get('category') == category]
    
    if priority:
        filtered_tickets = [ticket for ticket in filtered_tickets if ticket.get('priority') == priority]
    
    # Sort by created date (newest first)
    filtered_tickets = sorted(filtered_tickets, key=lambda x: _parse_ts(x.get('created_at', '')), reverse=True)
    
    return render_template('tickets.html',
                         tickets=filtered_tickets,
                         ticket_statuses=TICKET_STATUSES,
                         ticket_categories=TICKET_CATEGORIES,
                         priority_levels=PRIORITY_LEVELS,
                         search=search,
                         selected_status=status,
                         selected_category=category,
                         selected_priority=priority)

@app.route('/tickets/add', methods=['GET', 'POST'])
def add_ticket():
    """Submit a new support ticket"""
    if request.method == 'POST':
        # Validate required fields
        title = request.form.get('title')
        category = request.form.get('category')
        submitted_by = request.form.get('submitted_by')
        
        if not all([title, category, submitted_by]):
            flash('Please fill in all required fields.', 'error')
            return redirect(url_for('add_ticket'))

        now = datetime.now().strftime('%m-%d-%Y %H:%M:%S')
        conn = get_db()
        try:
            cur = conn.execute('''INSERT INTO tickets
                (title, description, category, priority, status, submitted_by, created_at, updated_at)
                VALUES (?,?,?,?,?,?,?,?)''', (
                title, request.form.get('description', ''), category,
                request.form.get('priority', 'Medium'), 'Open', submitted_by, now, now
            ))
            ticket_id = cur.lastrowid
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()
        flash(f'Ticket #{ticket_id} created successfully!', 'success')
        return redirect(url_for('tickets_list'))
    
    return render_template('add_ticket.html',
                         ticket_categories=TICKET_CATEGORIES,
                         priority_levels=PRIORITY_LEVELS)

@app.route('/tickets/view/<int:ticket_id>')
def view_ticket(ticket_id):
    """View ticket details"""
    conn = get_db()
    row = conn.execute('SELECT * FROM tickets WHERE id = ?', (ticket_id,)).fetchone()
    conn.close()
    if not row:
        flash('Ticket not found.', 'error')
        return redirect(url_for('tickets_list'))
    return render_template('view_ticket.html',
                         ticket=dict(row),
                         ticket_statuses=TICKET_STATUSES)

@app.route('/tickets/update/<int:ticket_id>', methods=['POST'])
def update_ticket_status(ticket_id):
    """Update ticket status"""
    new_status = request.form.get('status', '').strip()
    if new_status not in TICKET_STATUSES:
        flash('Invalid status value.', 'error')
        return redirect(url_for('view_ticket', ticket_id=ticket_id))
    conn = get_db()
    try:
        result = conn.execute('UPDATE tickets SET status=?, updated_at=? WHERE id=?',
                              (new_status, datetime.now().strftime('%m-%d-%Y %H:%M:%S'), ticket_id))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    if result.rowcount:
        flash(f'Ticket #{ticket_id} status updated to {new_status}!', 'success')
    else:
        flash('Ticket not found.', 'error')
    return redirect(url_for('view_ticket', ticket_id=ticket_id))

@app.route('/tickets/delete/<int:ticket_id>', methods=['POST'])
def delete_ticket(ticket_id):
    """Delete a ticket"""
    conn = get_db()
    result = conn.execute('DELETE FROM tickets WHERE id=?', (ticket_id,))
    conn.commit()
    conn.close()
    if result.rowcount:
        flash('Ticket deleted successfully!', 'success')
    else:
        flash('Ticket not found.', 'error')
    return redirect(url_for('tickets_list'))

CSV_COLUMNS = ['device_type', 'model', 'serial_number', 'assigned_user', 'password', 'purchase_date', 'condition', 'location_code', 'island', 'phone', 'notes']

def _parse_ts(ts):
    try:
        return datetime.strptime(ts, '%m-%d-%Y %H:%M:%S')
    except (ValueError, TypeError):
        return datetime.min

def _normalize_date(date_str):
    """Parse common date formats and return MM-DD-YYYY. Returns original string if unparseable."""
    if not date_str or not date_str.strip():
        return date_str
    s = date_str.strip()
    for fmt in (
        '%m-%d-%Y',    # already correct: 03-07-2030
        '%m/%d/%y',    # 3/7/70  →  2070 (forced to 2000s)
        '%m/%d/%Y',    # 3/7/2030
        '%Y-%m-%d',    # 2030-03-07  (ISO / Excel)
        '%m-%d-%y',    # 03-07-70  →  2070 (forced to 2000s)
        '%B %d, %Y',   # January 15, 2030
        '%b %d, %Y',   # Jan 15, 2030
        '%B %d %Y',    # January 15 2030
        '%b %d %Y',    # Jan 15 2030
    ):
        try:
            parsed = datetime.strptime(s, fmt)
            if '%y' in fmt and parsed.year < 2000:
                parsed = parsed.replace(year=parsed.year + 100)
            return parsed.strftime('%m-%d-%Y')
        except ValueError:
            continue
    return s  # unrecognised — keep as-is

def _island_for_location(location_code):
    """Return the island for a given location code.
    Checks locations table first, then falls back to existing inventory data."""
    conn = get_db()
    row = conn.execute('SELECT island FROM locations WHERE code = ?', (location_code,)).fetchone()
    if row and row['island']:
        conn.close()
        return row['island']
    inv_row = conn.execute(
        "SELECT island FROM inventory WHERE location_code = ? AND island != '' LIMIT 1",
        (location_code,)
    ).fetchone()
    conn.close()
    return inv_row['island'] if inv_row else ''

def _build_locations_by_island(inventory=None):
    if inventory is None:
        inventory = load_inventory()
    loc_to_island = {}
    # Start with manually stored locations as base
    for entry in load_locations():
        code = entry.get('code', '').strip()
        island = entry.get('island', '').strip()
        if code:
            loc_to_island[code] = island or 'Other'
    # Inventory data always overrides stored locations when it has a non-empty island
    for item in inventory:
        loc = item.get('location_code', '').strip()
        island = item.get('island', '').strip()
        if loc:
            if island:
                # Non-empty island from inventory always wins
                loc_to_island[loc] = island
            elif loc not in loc_to_island:
                # Only fall back to 'Other' if not already mapped
                loc_to_island[loc] = 'Other'
    result = {}
    for loc, island in sorted(loc_to_island.items()):
        result.setdefault(island, []).append(loc)
    return dict(sorted(result.items()))

def _get_all_islands(inventory=None):
    """Build full island list from stored islands + inventory data."""
    if inventory is None:
        inventory = load_inventory()
    islands = set(load_islands())
    for item in inventory:
        isl = item.get('island', '').strip()
        if isl:
            islands.add(isl)
    return sorted(islands)
CSV_REQUIRED = {'device_type', 'location_code'}

COLUMN_ALIASES = {
    'device type': 'device_type',
    'devicetype': 'device_type',
    'type': 'device_type',
    'computername/servicetag': 'serial_number',
    'computername': 'serial_number',
    'service tag': 'serial_number',
    'servicetag': 'serial_number',
    'serial': 'serial_number',
    'serial #': 'serial_number',
    'serial no': 'serial_number',
    'location': 'location_code',
    'location code': 'location_code',
    'username': 'assigned_user',
    'user': 'assigned_user',
    'user name': 'assigned_user',
    'assigned user': 'assigned_user',
    'notes:': 'notes',
    'note': 'notes',
    'purchase date': 'purchase_date',
    'warranty end date': 'purchase_date',
    'warranty': 'purchase_date',
}

def _normalize_col(name):
    name = str(name).strip().lower()
    return COLUMN_ALIASES.get(name, name)

def _xlsx_cell_str(val):
    if val is None:
        return ''
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    if hasattr(val, 'strftime'):
        return val.strftime('%m-%d-%Y')
    return str(val).strip()

def _process_rows(rows):
    """Import a list of dicts (with CSV_COLUMNS keys) into inventory. Returns (imported, skipped, skip_reasons)."""
    saved_types = load_device_types()
    original_type_count = len(saved_types)
    imported, skipped = 0, 0
    skip_reasons = {}
    new_devices = []

    for row in rows:
        missing = [f for f in CSV_REQUIRED if not str(row.get(f, '')).strip()]
        if missing:
            reason = f'missing: {", ".join(missing)}'
            skip_reasons[reason] = skip_reasons.get(reason, 0) + 1
            skipped += 1
            continue

        device_type = str(row.get('device_type', '')).strip()
        matched_type = next((dt for dt in saved_types if dt.lower() == device_type.lower()), None)
        if not matched_type:
            matched_type = DEVICE_TYPE_ALIASES.get(device_type.lower())
        device_type = matched_type or device_type

        if device_type and device_type not in saved_types:
            saved_types.append(device_type)

        condition = str(row.get('condition', '')).strip()
        if condition not in DEVICE_CONDITIONS:
            condition = 'Good'

        new_devices.append({
            'device_type': device_type,
            'model': str(row.get('model', '')).strip() or '-',
            'serial_number': str(row.get('serial_number', '')).strip() or 'n/a',
            'assigned_user': (str(row.get('assigned_user', '')).strip() or
                              str(row.get('contact', '')).strip()),
            'password': str(row.get('password', '')).strip(),
            'purchase_date': _normalize_date(str(row.get('purchase_date', '')).strip()),
            'condition': condition,
            'location_code': str(row.get('location_code', '')).strip(),
            'island': str(row.get('island', '')).strip() or _island_for_location(str(row.get('location_code', '')).strip()),
            'phone': str(row.get('phone', '')).strip(),
            'notes': str(row.get('notes', '')).strip(),
            'added_at': datetime.now().strftime('%m-%d-%Y %H:%M:%S'),
        })
        imported += 1

    if new_devices:
        conn = get_db()
        try:
            conn.execute('BEGIN')
            for d in new_devices:
                conn.execute('''INSERT INTO inventory
                    (device_type, model, serial_number, assigned_user, password,
                     purchase_date, condition, location_code, island, phone, notes,
                     archived, added_at, updated_at, source)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,0,?,?,?)''', (
                    d['device_type'], d['model'], d['serial_number'], d['assigned_user'],
                    d['password'], d['purchase_date'], d['condition'], d['location_code'],
                    d['island'], d['phone'], d['notes'], d['added_at'], '', 'imported'
                ))
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    if len(saved_types) > original_type_count:
        save_device_types(saved_types)

    return imported, skipped, skip_reasons


@app.route('/inventory/import', methods=['GET', 'POST'])
def import_inventory():
    if request.method == 'POST':
        # Phase 2: user selected a sheet from an xlsx
        if request.form.get('sheet_name'):
            sheet_name = request.form.get('sheet_name')
            temp_path = session.pop('xlsx_temp_path', '')

            if not temp_path or not os.path.exists(temp_path):
                flash('Upload session expired. Please upload the file again.', 'error')
                return redirect(url_for('import_inventory'))

            try:
                wb = openpyxl.load_workbook(temp_path)
                ws = wb[sheet_name]
                headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                header_map = {_normalize_col(h): i for i, h in enumerate(headers) if h is not None}

                rows = []
                for row in ws.iter_rows(min_row=2):
                    if all(cell.value is None for cell in row):
                        continue
                    rows.append({
                        col: (_xlsx_cell_str(row[header_map[col]].value) if col in header_map else '')
                        for col in CSV_COLUMNS + ['contact']
                    })
                wb.close()
            except Exception as e:
                flash(f'Failed to read file: {e}', 'error')
                return redirect(url_for('import_inventory'))
            finally:
                if os.path.exists(temp_path):
                    os.remove(temp_path)

            imported, skipped, skip_reasons = _process_rows(rows)

            if imported:
                log_history('Imported', {
                    'id': None, 'device_type': '', 'model': f'{imported} device(s) from "{sheet_name}"',
                    'serial_number': '', 'location_code': '', 'island': '', 'assigned_user': ''
                })
                msg = f'Successfully imported {imported} device(s) from "{sheet_name}".'
                if skipped:
                    reasons = '; '.join(f'{count}x {reason}' for reason, count in skip_reasons.items())
                    msg += f' Skipped {skipped} row(s): {reasons}.'
                flash(msg, 'success')
            else:
                detected = ', '.join(header_map.keys()) or 'none'
                reasons = '; '.join(f'{count}x {reason}' for reason, count in skip_reasons.items())
                flash(f'No devices imported from "{sheet_name}". {skipped} row(s) skipped — {reasons}. Columns detected: [{detected}].', 'error')

            return redirect(url_for('inventory_list'))

        # Phase 1: file upload
        file = request.files.get('csv_file')
        if not file or not file.filename:
            flash('Please select a file.', 'error')
            return redirect(url_for('import_inventory'))

        filename = file.filename.lower()

        if filename.endswith('.csv'):
            raw = file.stream.read()
            try:
                content = raw.decode('utf-8-sig')
            except UnicodeDecodeError:
                content = raw.decode('latin-1')
            stream = io.StringIO(content)
            rows = [{_normalize_col(k): v for k, v in row.items()} for row in csv.DictReader(stream)]
            imported, skipped, skip_reasons = _process_rows(rows)

            if imported:
                log_history('Imported', {
                    'id': None, 'device_type': '', 'model': f'{imported} device(s) from "{file.filename}"',
                    'serial_number': '', 'location_code': '', 'island': '', 'assigned_user': ''
                })
                msg = f'Successfully imported {imported} device(s).'
                if skipped:
                    reasons = '; '.join(f'{count}x {reason}' for reason, count in skip_reasons.items())
                    msg += f' Skipped {skipped} row(s): {reasons}.'
                flash(msg, 'success')
            else:
                reasons = '; '.join(f'{count}x {reason}' for reason, count in skip_reasons.items())
                flash(f'No devices imported. {skipped} row(s) skipped — {reasons}.', 'error')

            return redirect(url_for('inventory_list'))

        elif filename.endswith('.xlsx'):
            temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
            os.close(temp_fd)
            try:
                file.save(temp_path)
                wb = openpyxl.load_workbook(temp_path, read_only=True)
                sheets = wb.sheetnames
                wb.close()
            except Exception as e:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                flash(f'Failed to read Excel file: {e}', 'error')
                return redirect(url_for('import_inventory'))

            session['xlsx_temp_path'] = temp_path
            return render_template('import_inventory.html', sheets=sheets)

        else:
            flash('Please upload a .csv or .xlsx file.', 'error')
            return redirect(url_for('import_inventory'))

    return render_template('import_inventory.html')


@app.route('/inventory/import/template')
def download_csv_template():
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=CSV_COLUMNS)
    writer.writeheader()
    writer.writerow({
        'device_type': 'Laptop',
        'model': 'Dell XPS 15',
        'serial_number': 'SN123456',
        'assigned_user': 'Jane Doe',
        'purchase_date': '2024-01-15',
        'condition': 'Good',
        'location_code': '101',
        'notes': 'Example row — delete before importing'
    })
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=inventory_template.csv'}
    )


@app.route('/inventory/export')
def export_inventory():
    filter_type = request.args.get('filter', 'active')
    inventory = load_inventory()

    if filter_type == 'archived':
        devices = [d for d in inventory if d.get('archived')]
        filename = 'inventory_archived.xlsx'
    elif filter_type == 'all':
        devices = inventory
        filename = 'inventory_all.xlsx'
    else:
        devices = [d for d in inventory if not d.get('archived')]
        filename = 'inventory_active.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filter_type.capitalize()

    headers = [
        'ComputerName/ServiceTag', 'Location', 'Device Type', 'Model',
        'Username', 'Password', 'Island', 'Location Address w/zip',
        'Phone', 'Contact', 'Warranty End Date', 'Memory', 'CPU',
        'Operating Sys', 'NOTES:', 'Archived'
    ]
    ws.append(headers)

    # Bold header row
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for d in devices:
        ws.append([
            d.get('serial_number', ''),
            d.get('location_code', ''),
            d.get('device_type', ''),
            d.get('model', ''),
            d.get('assigned_user', ''),
            d.get('password', ''),
            d.get('island', ''),
            '',                               # Location Address w/zip — not stored
            d.get('phone', ''),
            '',                               # Contact — not stored separately
            d.get('purchase_date', ''),
            '',                               # Memory — not stored
            '',                               # CPU — not stored
            '',                               # Operating Sys — not stored
            d.get('notes', ''),
            'Yes' if d.get('archived') else 'No',
        ])

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


if __name__ == '__main__':
    init_data_storage()
    app.run(debug=True, host='0.0.0.0', port=5001)

